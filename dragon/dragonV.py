import cv2
import os
import json
import openpyxl as xl
import pandas as pd
from sklearn.preprocessing import MinMaxScaler
import numpy as np
POSE_2D_JOINT_NAME_LIST = [
        'Nose_x', 'Nose_y', 
        'Neck_x', 'Neck_y', 
        'RShoulder_x', 'RShoulder_y', 
        'RElbow_x', 'RElbow_y', 
        'RWrist_x', 'RWrist_y', 
        'LShoulder_x', 'LShoulder_y', 
        'LElbow_x', 'LElbow_y', 
        'LWrist_x', 'LWrist_y', 
        'MidHip_x', 'MidHip_y', 
        'RHip_x', 'RHip_y',
        'RKnee_x', 'RKnee_y',
        'RAnkle_x', 'RAnkle_y', 
        'LHip_x', 'LHip_y',
        'LKnee_x', 'LKnee_y', 
        'LAnkle_x', 'LAnkle_y',
        'REye_x', 'REye_y',
        'LEye_x', 'LEye_y', 
        'REar_x', 'REar_y', 
        'LEar_x', 'LEar_y', 
        'LBigToe_x', 'LBigToe_y', 
        'LSmallToe_x', 'LSmallToe_y', 
        'LHeel_x', 'LHeel_y',
        'RBigToe_x', 'RBigToe_y',
        'RSmallToe_x', 'RSmallToe_y', 
        'RHeel_x', 'RHeel_y'
]

def add_number(a:int, b:int) -> int:
    return a+b

def remove_confidence_from_keypoints_2d(p_keypoints):
    return [value for index, value in enumerate(p_keypoints, start=1) if index % 3!= 0]

def nomalize_data(frame_list):
    scalar = MinMaxScaler()
    nomalized_frame_list = scalar.fit_transform(frame_list).tolist()
    
    return nomalized_frame_list

def framelist2excel(frame_list, feature_list, save_path):
    wb = xl.Workbook()
    sheet = wb.active

    sheet.append(feature_list)
    for each_frame in frame_list:
        sheet.append(each_frame)
    
    wb.save(save_path)
    print(f'[save done] path : {save_path}\n')
    
def mark_pos(img, x, y, radius = 5, color = (0, 0, 255), thickness = 2):
    _x = round(x)
    _y = round(y)
    cv2.circle(img, (_x, _y), radius, color, thickness)

def get_total_frame(video_path):
    cap = cv2.VideoCapture(video_path)
    return int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

def get_jsons_list(folder_path):
    file_list = [f for f in sorted(os.listdir(folder_path)) if f.endswith('.json')]
    return file_list

def save_video2Img(video, path):
    pass

def get_all_frame_data_from_jsons_list(jsons_list:list, root_dir:str) -> list :
    #3차원 배열
    all_frame_data = []

    for now_json in jsons_list:
        with open(root_dir + now_json, 'r') as file:
            data = json.load(file)
        
        people = data['people']
        len_people = len(people)

        now_frame = []
        for idx in range(0, len_people):
            p = people[idx]
            p_keypoints_2d = p['pose_keypoints_2d']

            p_keypoints_2d = remove_confidence_from_keypoints_2d(p_keypoints_2d)
            now_frame.append(p_keypoints_2d)
        
        all_frame_data.append(now_frame)

    return all_frame_data

#this function is for only openpose json body25 model
def jsons2excel(json_dir_path, feature_list, key_index, file_name, save_path, json_key = "people"):
    #create xlsx file
    workbook = xl.Workbook()
    sheet = workbook.active
    
    #feature_name
    sheet.append(feature_list)

    #fetch json file list
    json_list = get_jsons_list(json_dir_path)
    
    #read json file sequentially
    for name in json_list:
        #read json file
        with open(json_dir_path + name, 'r') as file:
            data = json.load(file)

        people = data[json_key]
        p = people[key_index]
        p_keypoints = p["pose_keypoints_2d"]
        
        # except confidence value
        p_coord = [value for index, value in enumerate(p_keypoints, start=1) if index % 3!= 0]

        #save json data at each row
        sheet.append(p_coord)

    workbook.save(save_path + file_name)

'''
data_path : data type is only axcel available.
feature_list : feature_list of data. the type is list(str).
video_path : raw video path
save_video_path : processed video path where you want to save 
'''
#speed unit : ms
def mark_position_at_video(data_path, video_path, video_name, speed = 100):
    cap = cv2.VideoCapture(video_path)
    total_frame = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

    # read pos data
    wb = xl.load_workbook(data_path, data_only=True)
    #pos_sheet = wb.active
    pos_sheet = wb['frame']
    frame_pos_data_list = []
    
    for row in pos_sheet.iter_rows(values_only=True):
        frame_pos_data_list.append(list(row))

    num_label = len(frame_pos_data_list[0])
    
    #open video
    if not cap.isOpened():
        print("Error: Could not open video")
        return -1
    
    print(f"[read success ]{video_path}")

    while cap.isOpened():
        for frame_num in range(0, total_frame):
            ret, frame = cap.read()

            if not ret:
                break
            # processing area #
            for col in range(0, num_label, 2):
                x = frame_pos_data_list[frame_num + 1][col]
                y = frame_pos_data_list[frame_num + 1][col + 1]

                mark_pos(frame, x, y)
            ####################
            print(f"now frame : {frame_num}\n")
            cv2.imshow(video_name, frame)
            cv2.waitKey(speed)
        
        if cv2.waitKey(30) & 0xFF == 27:
            break

    cap.release()
    cv2.destroyAllWindows()

'''
all_frame_list : 각 프레임 값이 저장되어 있는 2차원 배열
'''
def play_marked_position_from_video(all_frame_list, video_path, video_name, speed=1):
    cap = cv2.VideoCapture(video_path)

    #open video
    if not cap.isOpened():
        print("Error: Could not open video")
        return -1
    
    print(f"[read success ]{video_path}")

    while cap.isOpened():
            ret, frame = cap.read()
            if not ret:
                break

            cur_frame = int(cap.get(cv2.CAP_PROP_POS_FRAMES)) - 1
            ####processing#######
            for idx in range(0, len(all_frame_list[cur_frame]), 2):
                x = all_frame_list[cur_frame][idx]
                y = all_frame_list[cur_frame][idx + 1]

                mark_pos(frame, x, y)
            ####################
            print(f'now frame : {cur_frame}')
            cv2.imshow(video_name, frame)
            cv2.waitKey(speed)
            if cv2.waitKey(30) & 0xFF == 27:
                break

    cap.release()
    cv2.destroyAllWindows()


def xlsx2data(data_path, sheet_name, start_row = 1):
    wb = xl.load_workbook(data_path, data_only=True)
    sheet = wb[sheet_name]
    frame_data_list = []

    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        frame_data_list.append(list(row))
    
    return frame_data_list

def xlsx2data(data_path, sheet_name, start_row = 1):
    wb = xl.load_workbook(data_path, data_only=True)
    sheet = wb[sheet_name]
    frame_data_list = []

    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        frame_data_list.append(list(row))
    

    return frame_data_list
'''
openpose로 joint position estimation을 했을 때,
영상에 여러명의 사람이 나오는 경우 json파일에서 2d joint pos의 idx가 바뀌는 경우가
발생할 수 있음.

이런 경우에 사용할 수 있는 함수.
제일 가능성이 높은 다음 인덱스를 반환

방식)
현재 프레임과 다음 프레임의 각 joint pos의 거리의 합을 인덱스별로 계산해서,
제일 작은 값을 갖는 인덱스를 선택.
'''
def get_estimation_next_idx(base_idx : int, num_now_frame : int, json_dir_path : str) -> int:

    json_list = get_jsons_list(json_dir_path)
        ######기준이 되는 frame 좌표 가져오기#######
    with open(json_dir_path + json_list[num_now_frame], 'r') as file:
        data = json.load(file)
    people = data['people']
    #people idx 지정
    p = people[base_idx]
    #pose_keypoints_2d 가져오기
    p_keypoints = p['pose_keypoints_2d']
    #confidence 값 제거
    now_frame = remove_confidence_from_keypoints_2d(p_keypoints)
    ####################################

    ####다음 frame 좌표 가져오기##############
    nxt_frame_json = json_list[num_now_frame + 1]
    candidate_frame = []
    with open(json_dir_path + nxt_frame_json, 'r') as file:
        data = json.load(file)
    people = data['people']
    num_people = len(people)

    for idx in range(0, num_people):
        p = people[idx]
        p_keypoints = p['pose_keypoints_2d']
        p_keypoints_2d = remove_confidence_from_keypoints_2d(p_keypoints)
        candidate_frame.append(p_keypoints_2d)
    ######################################
    
    ### 거리 계산 log(norm2)###
    distnace_list = []
    for nxt_frame in candidate_frame:
        total_loss = 0
        for i in range(0, len(now_frame), 2):
            x = now_frame[i]
            y = now_frame[i+1]
            nx = nxt_frame[i]
            ny = nxt_frame[i+1]

            total_loss += (x-nx)**2 + (y-ny)**2
        
        distnace_list.append(np.round(np.log(total_loss), 5))
    ################################
    return distnace_list.index((min(distnace_list)))
    
def get_poskeypoints2d_from_json(file_path:str, people_idx:int)-> list:
    with open(file_path, 'r') as file:
        data = json.load(file)
        people = data['people']
        
        p = people[people_idx]
        p_keypoints = p["pose_keypoints_2d"]

        return [value for index, value in enumerate(p_keypoints, start=1) if index % 3!= 0]

def play_video(video_path, video_name):
    cap = cv2.VideoCapture(video_path)

    if not cap.isOpened():
        print(f'[Error: could not open video]:{video_path}')
        return -1
    
    print(f"[read success]{video_path}")
    while cap.isOpened():
        ret, frame = cap.read()
        if not ret:
            break
    
        #현재 프레임 번호 출력
        current_frame = int(cap.get(cv2.CAP_PROP_POS_FRAMES)) - 1
        print(f'current_frame : {current_frame}')

        cv2.imshow(video_name, frame)        
        if cv2.waitKey(25) & 0xFF == 27:  # esc 키의 ASCII 코드는 27
            break
       
    cap.release()
    cv2.destroyAllWindows()

def show_specific_video_frame(video_path, video_name : str, frame_num = 1):
    cap = cv2.VideoCapture(video_path)
    frame_number = frame_num
    cap.set(cv2.CAP_PROP_POS_FRAMES, frame_number)
    ret, frame = cap.read()
    # 비디오 파일 닫기
    cap.release()

    # 프레임 표시
    cv2.imshow(video_name, frame)
    cv2.waitKey(0)
    cv2.destroyAllWindows()

def get_specific_video_frame(video_path:str, frame_num = 0) -> np.ndarray:
    cap = cv2.VideoCapture(video_path)
    frame_number = frame_num
    cap.set(cv2.CAP_PROP_POS_FRAMES, frame_number)
    ret, frame = cap.read()
    # 비디오 파일 닫기
    cap.release()
    
    if ret == False:
        return False
    
    return frame

if __name__=="__main__":
    print(os.__version__)
